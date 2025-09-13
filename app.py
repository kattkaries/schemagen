import io
import math
import random
import time
import re
from collections import Counter
from datetime import date

import openpyxl
import plotly.express as px
import streamlit as st
from supabase import Client, create_client

# --- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="Schemagenerator",
    page_icon="📅",
    layout="centered",
)

# --- CSS FOR STYLING THE MULTISELECT WIDGET ---
# This CSS targets the selection "tags" in the first multiselect widget on the page,
# giving them a custom background color for better visual distinction.
st.markdown("""
<style>
    /* Target the container for the first multiselect widget */
    div[data-testid="stMultiSelect"]:first-of-type {
        /* This selector is for scoping purposes; no specific style is needed here. */
    }

    /* Style the selected "tags" within the first multiselect widget */
    div[data-testid="stMultiSelect"]:first-of-type [data-baseweb="tag"] {
        background-color: #2E8B57; /* SeaGreen */
        border-radius: 0.5rem;   /* Optional: for rounded corners */
    }

    /* Improve contrast by making the text and 'x' button white */
    div[data-testid="stMultiSelect"]:first-of-type [data-baseweb="tag"] span,
    div[data-testid="stMultiSelect"]:first-of-type [data-baseweb="tag"] span[role="button"] {
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

# --- SWEDISH TRANSLATION SETUP ---
SWEDISH_DAYS = {
    "Monday": "Måndag",
    "Tuesday": "Tisdag",
    "Wednesday": "Onsdag",
    "Thursday": "Torsdag",
    "Friday": "Fredag",
}

# --- SESSION STATE FLAGS ---
if "confirm_delete" not in st.session_state:
    st.session_state.confirm_delete = False

# --- SUPABASE CLIENT ---
try:
    supabase_url = st.secrets["SUPABASE_URL"]
    supabase_key = st.secrets["SUPABASE_KEY"]
    supabase: Client = create_client(supabase_url, supabase_key)
except Exception:
    st.error("Could not connect to the database. Please check your Supabase credentials.")
    st.stop()


# --- DATA FETCHING AND CACHING ---
@st.cache_data(ttl=600)
def fetch_all_data():
    """
    Fetches MDK assignments, Screen/MR sessions, and employee work rates.
    """
    # MDK
    try:
        mdk_response = supabase.table("mdk_assignments").select("employee, week, day").execute()
        mdk_data = mdk_response.data if mdk_response.data else []
    except Exception:
        mdk_data = []

    # Work rates
    try:
        work_rate_response = supabase.table("work_rates").select("employee, rate").execute()
        work_rate_data = work_rate_response.data if work_rate_response.data else []
    except Exception:
        work_rate_data = []

    # Screen/MR sessions
    try:
        screen_mr_response = supabase.table("screen_mr_sessions").select("employee, week, day, block").execute()
        screen_mr_data = screen_mr_response.data if screen_mr_response.data else []
    except Exception:
        screen_mr_data = []

    return mdk_data, work_rate_data, screen_mr_data


# Cached data (refresh via st.cache_data.clear() when mutated)
all_mdk_assignments, db_work_rates_list, all_screen_mr_sessions = fetch_all_data()

# --- PRE-POPULATED DATA & CONSTANTS ---
PRE_POP_EMPLOYEES = ["AH", "LS", "DS", "KL", "TH", "LAO", "AL", "HS", "AG", "CB", "NC"]
PRE_UNAVAILABLE = {
    "Monday": ["DS", "HS", "LS"],
    "Tuesday": ["LAO", "CB", "HS", "LS"],
    "Wednesday": ["DS", "AH", "CB", "KL"],
    "Thursday": ["CB", "KL", "NC"],
    "Friday": ["CB", "AL", "KL"],
}
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# --- Screen/MR configuration ---
SCREEN_MR_PER_BLOCK = 1           # number of Screen/MR assignees per morning/afternoon
SCREEN_MR_WEEKLY_CAP = 1          # soft cap per person/week


# --- HELPERS: Weighted selection with weekly cap ---
def _unique_weighted_choices(candidates, weight_lookup, k):
    """
    Choose k unique items from candidates using their weights (probabilities ∝ work rate).
    Falls back to near-uniform if all weights are zero.
    """
    if k <= 0 or not candidates:
        return []
    picks = []
    pool = list(candidates)
    while pool and len(picks) < k:
        weights = [max(0.001, float(weight_lookup.get(c, 0))) for c in pool]
        total = sum(weights)
        probs = None if total <= 0 else [w / total for w in weights]
        chosen = random.choices(pool, weights=probs, k=1)[0]
        picks.append(chosen)
        pool.remove(chosen)
    return picks


def weighted_sample_with_cap(candidates, weight_lookup, k, weekly_counts, cap):
    """
    Prefer people under the weekly cap; if not enough candidates under-cap, fill from over-cap pool.
    """
    if k <= 0 or not candidates:
        return []
    under_cap = [c for c in candidates if weekly_counts.get(c, 0) < cap]
    if len(under_cap) >= k:
        return _unique_weighted_choices(under_cap, weight_lookup, k)
    picks = under_cap[:]
    remaining_k = k - len(picks)
    over_cap_pool = [c for c in candidates if c not in picks]
    if remaining_k > 0 and over_cap_pool:
        picks.extend(_unique_weighted_choices(over_cap_pool, weight_lookup, remaining_k))
    return picks


# --- UI: TITLE AND INSTRUCTIONS ---
st.title("📅 Schemagenerator för världens bästa enhet!")
st.info(
    "**💡 Användarinstruktioner:**\n\n"
    "1. **Ange frånvaro:** Markera medarbetare som är frånvarande hela veckan eller specifika dagar.\n"
    "2. **Granska & justera:** Kontrollera MDK-historiken och arbetstiden. Spara eventuella ändringar i arbetstid.\n"
    "3. **Generera:** Klicka på '✨ Generera Schema' för att skapa ett nytt veckoschema.\n"
    "4. **Ladda hem:** Klicka på '📥 Ladda ner schemat' för ta hem det."
)

# --- UI: EMPLOYEE AVAILABILITY ---
available_week = st.multiselect(
    "🙋 Initialer för samtliga medarbetare denna vecka",
    options=PRE_POP_EMPLOYEES,
    default=PRE_POP_EMPLOYEES,
)

unavailable_whole_week = st.multiselect(
    "🏖️ Initialer för medarbetare som är otillgängliga hela veckan",
    options=available_week,
    default=[],
)

# Filter out employees who are unavailable for the entire week.
available_employees = [emp for emp in available_week if emp not in unavailable_whole_week]

# Per-day unavailability
with st.expander("📅 Ange otillgänglighet per dag", expanded=True):
    unavailable_per_day = {}
    for day in DAYS:
        default_values = [emp for emp in PRE_UNAVAILABLE.get(day, []) if emp in available_employees]
        unavailable_per_day[day] = st.multiselect(
            f"Frånvarande på {SWEDISH_DAYS[day]}",
            options=available_employees,
            default=default_values,
            key=f"unavail_{day}",
        )

# --- UI: MDK DISTRIBUTION CHART (historik) ---
with st.expander("📊 MDK-fördelning (historik)"):
    if all_mdk_assignments:
        mdk_counts = Counter(assignment["employee"] for assignment in all_mdk_assignments)
        sorted_items = sorted(mdk_counts.items(), key=lambda x: x[1], reverse=True)
        employees, counts = zip(*sorted_items)
        fig = px.bar(
            x=employees,
            y=counts,
            labels={"x": "Medarbetare", "y": "Antal MDK"},
            title="MDK-fördelning (baserat på sparad historik)",
            color=counts,
            color_continuous_scale="blugrn",
        )
        fig.update_layout(xaxis={"categoryorder": "total descending"})
        fig.update_coloraxes(showscale=False)
        fig.update_yaxes(dtick=1)
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Inga MDK-uppdrag finns i historiken ännu.")

# --- NEW UI: Screen/MR DISTRIBUTION CHART (historik) ---
with st.expander("📊 Screen/MR‑fördelning (historik)"):
    if all_screen_mr_sessions:
        weeks_available = sorted({row["week"] for row in all_screen_mr_sessions}, reverse=True)
        selected_weeks = st.multiselect(
            "Filtrera veckor (standard: alla)",
            options=weeks_available,
            default=weeks_available,
        )
        view_mode = st.radio(
            "Visa som",
            options=["Total", "Per block (stacked)"],
            horizontal=True,
            key="screenmr_view_mode",
        )

        filtered = [
            row for row in all_screen_mr_sessions
            if row.get("employee") in PRE_POP_EMPLOYEES
            and (not selected_weeks or row.get("week") in selected_weeks)
        ]

        if not filtered:
            st.info("Ingen Screen/MR‑data matchar det valda filtert.")
        else:
            if view_mode == "Total":
                sm_counts = Counter(row["employee"] for row in filtered)
                data = [{"employee": e, "count": c} for e, c in sm_counts.items()]
                data.sort(key=lambda x: x["count"], reverse=True)

                fig_sm = px.bar(
                    data,
                    x="employee",
                    y="count",
                    labels={"employee": "Medarbetare", "count": "Antal Screen/MR‑pass"},
                    title="Screen/MR‑fördelning (baserat på uppladdad historik)",
                    color="count",
                    color_continuous_scale="teal",
                )
                fig_sm.update_layout(xaxis={"categoryorder": "total descending"})
                fig_sm.update_coloraxes(showscale=False)
                fig_sm.update_yaxes(dtick=1)
                st.plotly_chart(fig_sm, use_container_width=True)

                # CSV export
                csv = "employee,count\n" + "\n".join(f"{d['employee']},{d['count']}" for d in data)
                st.download_button(
                    "⬇️ Ladda ner som CSV",
                    data=csv.encode("utf-8"),
                    file_name="screen_mr_distribution.csv",
                    mime="text/csv",
                )

            else:  # Per block (stacked)
                # Aggregate by (employee, block)
                counts_by_pair = Counter((row["employee"], row.get("block", "unknown")) for row in filtered)
                blocks = ["morning", "afternoon"]
                # Build stacked chart data
                bar_data = []
                for (emp, blk), c in counts_by_pair.items():
                    if blk not in blocks:
                        continue
                    bar_data.append({"employee": emp, "block": "FM" if blk == "morning" else "EM", "count": c})
                # Ensure deterministic ordering
                bar_data.sort(key=lambda x: (x["employee"], x["block"]))

                fig_stacked = px.bar(
                    bar_data,
                    x="employee",
                    y="count",
                    color="block",
                    barmode="stack",
                    labels={"employee": "Medarbetare", "count": "Antal pass", "block": "Block"},
                    title="Screen/MR‑fördelning per block (baserat på uppladdad historik)",
                    color_discrete_map={"FM": "#1f77b4", "EM": "#ff7f0e"},
                )
                fig_stacked.update_layout(xaxis={"categoryorder": "total descending"})
                fig_stacked.update_yaxes(dtick=1)
                st.plotly_chart(fig_stacked, use_container_width=True)

                # CSV export
                csv2 = "employee,block,count\n" + "\n".join(f"{d['employee']},{d['block']},{d['count']}" for d in bar_data)
                st.download_button(
                    "⬇️ Ladda ner som CSV",
                    data=csv2.encode("utf-8"),
                    file_name="screen_mr_distribution_by_block.csv",
                    mime="text/csv",
                )
    else:
        st.info("Ingen Screen/MR‑historik uppladdad ännu.")

# --- UI: HISTORICAL SCHEDULES (Senaste 8 veckorna) + CLEAR HISTORY (inside expander) ---
current_week = date.today().isocalendar()[1]
with st.expander("📝 Historiska scheman (Senaste 8 veckorna)"):
    try:
        bucket_files = supabase.storage.from_("schedules").list()
        file_names = {f["name"] for f in bucket_files} if bucket_files else set()
    except Exception as e:
        file_names = set()
        st.warning(f"Kunde inte hämta fillistan från lagringen: {e}")

    for i in range(1, 9):
        week = current_week - i
        file_name = f"week_{week}.xlsx"
        status_emoji = "✅" if file_name in file_names else "❌"
        st.write(f"**Vecka {week}:** {status_emoji} {'Uppladdat' if file_name in file_names else 'Ej uppladdat'}")

        uploader = st.file_uploader(
            f"Ladda upp/ersätt schema för vecka {week}",
            type="xlsx",
            key=f"hist_{week}",
        )

        if uploader:
            with st.spinner(f"Laddar upp och bearbetar {file_name}..."):
                file_content = uploader.getvalue()
                supabase.storage.from_("schedules").upload(file_name, file_content, {"upsert": "true"})
                st.success(f"Laddade upp {file_name}")
                time.sleep(1)

                # Parse the uploaded file to extract MDK and Screen/MR history
                downloaded = supabase.storage.from_("schedules").download(file_name)
                if downloaded:
                    wb = openpyxl.load_workbook(io.BytesIO(downloaded))
                    sheet = wb["Blad1"] if "Blad1" in wb.sheetnames else wb.active

                    # Mappings used in the template
                    mdk_cols = {"Monday": "D", "Tuesday": "H", "Thursday": "P"}
                    screen_cols = {"Monday": "C", "Tuesday": "G", "Wednesday": "K", "Thursday": "O", "Friday": "S"}

                    # --- Parse MDK (single cell per MDK day @ row 3) ---
                    parsed_mdk = []
                    for d, col in mdk_cols.items():
                        col_idx = openpyxl.utils.column_index_from_string(col)
                        cell_value = sheet.cell(row=3, column=col_idx).value
                        if cell_value and isinstance(cell_value, str) and cell_value.strip() in PRE_POP_EMPLOYEES:
                            parsed_mdk.append({"week": week, "day": d, "employee": cell_value.strip()})

                    # --- Parse Screen/MR (morning row 3, afternoon row 14; may contain "A/B") ---
                    def parse_initials(value):
                        if not value:
                            return []
                        s = str(value)
                        tokens = [t.strip() for t in re.split(r'[/,|]+', s) if t.strip()]
                        return [t for t in tokens if t in PRE_POP_EMPLOYEES]

                    parsed_screen_mr = []
                    for d in DAYS:
                        # Morning
                        m_col_idx = openpyxl.utils.column_index_from_string(screen_cols[d])
                        m_val = sheet.cell(row=3, column=m_col_idx).value
                        for emp in parse_initials(m_val):
                            parsed_screen_mr.append({"week": week, "day": d, "block": "morning", "employee": emp})
                        # Afternoon
                        a_val = sheet.cell(row=14, column=m_col_idx).value
                        for emp in parse_initials(a_val):
                            parsed_screen_mr.append({"week": week, "day": d, "block": "afternoon", "employee": emp})

                    # --- Save parsed data to Supabase (replace only THIS week) ---
                    try:
                        if parsed_mdk:
                            supabase.table("mdk_assignments").delete().eq("week", week).execute()
                            supabase.table("mdk_assignments").upsert(parsed_mdk).execute()

                        if parsed_screen_mr:
                            supabase.table("screen_mr_sessions").delete().eq("week", week).execute()
                            supabase.table("screen_mr_sessions").upsert(parsed_screen_mr).execute()

                        if parsed_mdk or parsed_screen_mr:
                            st.success(
                                f"Sparade historik för vecka {week}: "
                                f"{len(parsed_mdk)} MDK och {len(parsed_screen_mr)} Screen/MR."
                            )
                            st.cache_data.clear()
                            time.sleep(0.8)
                            st.rerun()
                        else:
                            st.info(f"Inga giltiga MDK eller Screen/MR-initialer hittades i filen för vecka {week}.")
                    except Exception as e:
                        st.error(f"Misslyckades med att spara historik: {e}")

    # --- CLEAR MDK HISTORY (inside this expander) ---
    st.markdown("---")
    st.error("Radering av historik kan inte ångras.")
    if st.button("🗑️ Rensa all MDK-historik", key="btn_clear_mdk"):
        st.session_state.confirm_delete = True

    if st.session_state.confirm_delete:
        st.warning("**Är du helt säker på att du vill radera ALL MDK-historik?**")
        col1, col2, _ = st.columns([1.5, 1, 4])
        with col1:
            if st.button("Ja, radera all historik", type="primary", key="btn_confirm_clear"):
                try:
                    supabase.table("mdk_assignments").delete().neq("week", -1).execute()
                    st.success("All MDK-historik har raderats.")
                    st.session_state.confirm_delete = False
                    st.cache_data.clear()
                    time.sleep(2)
                    st.rerun()
                except Exception as e:
                    st.error(f"Ett fel uppstod vid radering: {e}")
        with col2:
            if st.button("Avbryt", key="btn_cancel_clear"):
                st.session_state.confirm_delete = False
                st.rerun()

# --- UI: WORK RATES ---
db_work_rates = {row["employee"]: row["rate"] for row in db_work_rates_list}
default_work_rates = {emp: 100 for emp in PRE_POP_EMPLOYEES}
work_rates_initial = {**default_work_rates, **db_work_rates}

if "work_rates" not in st.session_state:
    st.session_state["work_rates"] = work_rates_initial.copy()

with st.expander("💼 Klinisk arbetstid per medarbetare (%)"):
    col1, col2 = st.columns(2)
    sorted_employees = sorted(PRE_POP_EMPLOYEES)
    midpoint = math.ceil(len(sorted_employees) / 2)

    for i, emp in enumerate(sorted_employees):
        target_col = col1 if i < midpoint else col2
        key = f"rate_{emp}"
        value_from_state = int(st.session_state["work_rates"].get(emp, 100))
        st.session_state["work_rates"][emp] = target_col.number_input(
            f"{emp} arbetstid",
            min_value=0,
            max_value=100,
            value=value_from_state,
            step=5,
            key=key,
            help=f"Ange den procentuella kliniska arbetstiden för {emp}.",
        )

    if st.button("💾 Spara arbetstid till databasen"):
        try:
            records_to_save = [
                {"employee": emp, "rate": st.session_state["work_rates"][emp]}
                for emp in PRE_POP_EMPLOYEES
            ]
            supabase.table("work_rates").upsert(records_to_save).execute()
            st.success("Arbetstid sparad!")
            st.cache_data.clear()
            time.sleep(1)
            st.rerun()
        except Exception as e:
            st.error(f"Ett fel uppstod vid sparning: {e}")

# Use the latest work rates from session state for generation.
work_rates = st.session_state["work_rates"]

# --- UI: GENERATE SCHEDULE ---
if st.button("✨ Generera Schema", type="primary"):
    with st.spinner("Tänker, slumpar och skapar... ett ögonblick..."):
        # --- MDK ASSIGNMENT LOGIC (no DB writes here) ---
        mdk_days = ["Monday", "Tuesday", "Thursday"]
        mdk_assignments = {}

        # Pre-calculate historical MDK counts (from uploaded history only)
        mdk_history_counts = Counter(a["employee"] for a in all_mdk_assignments)
        assigned_this_week = Counter()

        for day in mdk_days:
            # Who can take MDK on this day (exclude AL from MDK)
            avail_for_day = [
                emp
                for emp in available_employees
                if emp not in unavailable_per_day.get(day, [])
                and work_rates.get(emp, 0) > 0
                and emp != "AL"  # EXCLUDE AL from any MDK assignment
            ]
            if not avail_for_day:
                st.warning(f"Inga tillgängliga medarbetare för MDK/lunch på {SWEDISH_DAYS[day]}")
                continue

            # Score: lower is better
            scores = {}
            for emp in avail_for_day:
                history_count = mdk_history_counts.get(emp, 0)
                rate_factor = work_rates.get(emp, 100) / 100.0
                this_week_penalty = assigned_this_week[emp] * 10
                score = (history_count / rate_factor if rate_factor > 0 else float("inf")) + this_week_penalty
                scores[emp] = score

            chosen = min(scores, key=scores.get)
            mdk_assignments[day] = chosen
            assigned_this_week[chosen] += 1

        # --- SCHEDULE POPULATION LOGIC ---
        try:
            wb = openpyxl.load_workbook("template.xlsx")
            sheet = wb["Blad1"]
        except FileNotFoundError:
            st.error("`template.xlsx` hittades inte. Se till att filen ligger i samma mapp som appen.")
            st.stop()

        current_week = date.today().isocalendar()[1]
        sheet["A1"] = f"v.{current_week}"

        # Columns in template
        klin_cols = {"Monday": "B", "Tuesday": "F", "Wednesday": "J", "Thursday": "N", "Friday": "R"}
        screen_cols = {"Monday": "C", "Tuesday": "G", "Wednesday": "K", "Thursday": "O", "Friday": "S"}
        mdk_cols = {"Monday": "D", "Tuesday": "H", "Thursday": "P"}
        lunchvakt_col = {"Wednesday": "L"}

        lab_rows = {
            "morning1": {"LAB 3": 4, "LAB 6": 5, "LAB 9": 6, "LAB 10": 7},
            "morning2": {"LAB 3": 9, "LAB 6": 10, "LAB 9": 11, "LAB 10": 12},
            "afternoon1": {"LAB 3": 14, "LAB 6": 15, "LAB 9": 16, "LAB 10": 17},
        }
        labs = list(lab_rows["morning1"].keys())

        # Counters for fairness & weekly cap
        screen_mr_counts = Counter()       # for intra-week balancing lab vs. screen
        screen_mr_week_counts = Counter()  # for enforcing ~once/week cap on Screen/MR

        for day in DAYS:
            # Who is available today
            avail_day = [
                emp
                for emp in available_employees
                if emp not in unavailable_per_day.get(day, [])
            ]
            mdk = mdk_assignments.get(day)

            # Determine MDK behavior
            is_full_day_mdk = (day in ["Tuesday", "Thursday"]) and bool(mdk)  # Tue/Thu: full day MDK
            is_monday_mdk_morning = (day == "Monday") and bool(mdk)           # Mon: morning MDK

            # ============================
            # MORNING: FILL LABS FIRST
            # ============================
            lab_eligible_morning = [
                emp for emp in avail_day
                if not (is_full_day_mdk and emp == mdk)         # exclude Tue/Thu MDK entirely
                and not (is_monday_mdk_morning and emp == mdk)   # exclude Mon MDK from morning only
            ]
            lab_priority_candidates = sorted(
                lab_eligible_morning,
                key=lambda emp: screen_mr_counts.get(emp, 0),
                reverse=True,
            )
            morning_lab_slots = min(4, len(lab_priority_candidates))
            lab_people_morning = lab_priority_candidates[:morning_lab_slots]

            # Assign morning labs to template
            random.shuffle(labs)
            morning_assign = dict(zip(lab_people_morning, labs))

            klin_col = klin_cols[day]
            for p, l in morning_assign.items():
                sheet[f"{klin_col}{lab_rows['morning1'][l]}"] = p
                sheet[f"{klin_col}{lab_rows['morning2'][l]}"] = p

            # After labs, assign Screen/MR from the remainder only
            screen_pool_morning = [
                emp for emp in avail_day
                if emp not in lab_people_morning
                and work_rates.get(emp, 0) > 0
                and not (is_full_day_mdk and emp == mdk)
                and not (is_monday_mdk_morning and emp == mdk)
            ]

            if len(screen_pool_morning) > 0:
                screen_mr_morning = weighted_sample_with_cap(
                    candidates=screen_pool_morning,
                    weight_lookup=work_rates,
                    k=SCREEN_MR_PER_BLOCK,
                    weekly_counts=screen_mr_week_counts,
                    cap=SCREEN_MR_WEEKLY_CAP,
                )
            else:
                screen_mr_morning = []

            for s in screen_mr_morning:
                screen_mr_week_counts[s] += 1
                screen_mr_counts[s] += 1

            sheet[f"{screen_cols[day]}3"] = "/".join(screen_mr_morning) if screen_mr_morning else ""

            # ============================
            # AFTERNOON (not Friday): FILL LABS FIRST
            # ============================
            if day != "Friday":
                available_for_afternoon = [
                    emp for emp in avail_day
                    if not (is_full_day_mdk and emp == mdk)   # Tue/Thu MDK excluded all day
                ]
                afternoon_lab_slots = min(4, len(available_for_afternoon))

                preferred_candidates = [emp for emp in (screen_mr_morning or []) if emp in available_for_afternoon]
                other_candidates = [emp for emp in available_for_afternoon if emp not in preferred_candidates]
                combined_candidates = preferred_candidates + other_candidates
                lab_people_afternoon = combined_candidates[:afternoon_lab_slots]

                # Try to avoid same lab morning vs afternoon for the same person
                afternoon_labs = labs[:]
                afternoon_assign = {}
                for _ in range(10):
                    random.shuffle(afternoon_labs)
                    afternoon_assign = dict(zip(lab_people_afternoon, afternoon_labs))
                    if all(
                        afternoon_assign.get(p) != morning_assign.get(p)
                        for p in afternoon_assign
                        if p in morning_assign
                    ):
                        break

                for p, l in afternoon_assign.items():
                    sheet[f"{klin_col}{lab_rows['afternoon1'][l]}"] = p

                # After labs, assign Screen/MR if anyone remains
                afternoon_screen_pool = [
                    emp for emp in available_for_afternoon
                    if emp not in lab_people_afternoon
                    and work_rates.get(emp, 0) > 0
                ]
                if len(afternoon_screen_pool) > 0:
                    screen_mr_afternoon = weighted_sample_with_cap(
                        candidates=afternoon_screen_pool,
                        weight_lookup=work_rates,
                        k=SCREEN_MR_PER_BLOCK,
                        weekly_counts=screen_mr_week_counts,
                        cap=SCREEN_MR_WEEKLY_CAP,
                    )
                else:
                    screen_mr_afternoon = []

                for s in screen_mr_afternoon:
                    screen_mr_week_counts[s] += 1
                    screen_mr_counts[s] += 1

                sheet[f"{screen_cols[day]}14"] = "/".join(screen_mr_afternoon) if screen_mr_afternoon else ""

            # --- MDK & Lunch Guard (Wednesday only) ---
            if mdk:
                if day in mdk_cols:
                    sheet[f"{mdk_cols[day]}3"] = mdk
            elif day == "Wednesday":
                # Lunch guard: AL allowed normally; prefer non-lab morning people
                lunch_candidates = [p for p in avail_day if p not in lab_people_morning] or avail_day
                sheet[f"{lunchvakt_col['Wednesday']}3"] = random.choice(lunch_candidates) if lunch_candidates else ""

        # --- EXCEL DOWNLOAD (no DB writes here) ---
        output_file = io.BytesIO()
        wb.save(output_file)
        output_file.seek(0)

        st.success("✅ Schemat har genererats!")
        st.download_button(
            label="📥 Ladda ner schemat (.xlsx)",
            data=output_file,
            file_name=f"veckoschema_v{current_week}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
